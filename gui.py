import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
from typing import List
import threading
import openpyxl
from openpyxl.styles import PatternFill, Font
from equipment import format_game_date
from collections import defaultdict
from eq_definitions import EQ_TYPE
from state_defines import STATES

class AnalyzerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("HOI4 Save Analyzer")
        self.root.geometry("800x600")
        
        # File type selection
        self.file_type = tk.StringVar(value="hoi4")
        self.setup_file_type_selector()
        
        # File list
        self.setup_file_list()
        
        # Buttons
        self.setup_buttons()
        
        # Progress area
        self.setup_progress_area()
        
        # Selected files
        self.selected_files = []
        
    def setup_file_type_selector(self):
        frame = ttk.LabelFrame(self.root, text="File Type", padding="5")
        frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Radiobutton(frame, text="HOI4 Saves (.hoi4)", 
                       variable=self.file_type, value="hoi4").pack(side=tk.LEFT)
        ttk.Radiobutton(frame, text="Melted Saves (.txt)", 
                       variable=self.file_type, value="txt").pack(side=tk.LEFT)
        ttk.Radiobutton(frame, text="JSON Files (.json)", 
                       variable=self.file_type, value="json").pack(side=tk.LEFT)
        
        # Add country tag input and button
        tag_frame = ttk.Frame(frame)
        tag_frame.pack(side=tk.RIGHT, padx=5)
        
        ttk.Label(tag_frame, text="Country TAG:").pack(side=tk.LEFT)
        self.country_tag = ttk.Entry(tag_frame, width=10)
        self.country_tag.pack(side=tk.LEFT, padx=2)
        ttk.Button(tag_frame, text="Generate Report", 
                  command=self.generate_country_report).pack(side=tk.LEFT, padx=2)
                       
    def setup_file_list(self):
        frame = ttk.LabelFrame(self.root, text="Selected Files", padding="5")
        frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Add listbox
        self.file_listbox = tk.Listbox(frame, yscrollcommand=scrollbar.set)
        self.file_listbox.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.file_listbox.yview)
        
    def setup_buttons(self):
        frame = ttk.Frame(self.root)
        frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(frame, text="Add Files", command=self.add_files).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame, text="Remove Selected", command=self.remove_selected).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame, text="Clear All", command=self.clear_files).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame, text="Generate General Report", command=self.process_files).pack(side=tk.RIGHT, padx=5)
        
    def setup_progress_area(self):
        frame = ttk.LabelFrame(self.root, text="Progress", padding="5")
        frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.progress_var = tk.StringVar(value="Ready")
        ttk.Label(frame, textvariable=self.progress_var).pack(fill=tk.X)
        
        self.progress_bar = ttk.Progressbar(frame, mode='determinate')
        self.progress_bar.pack(fill=tk.X, pady=(5,0))
        
    def add_files(self):
        filetypes = []
        if self.file_type.get() == "hoi4":
            filetypes = [("HOI4 Save Files", "*.hoi4")]
        elif self.file_type.get() == "txt":
            filetypes = [("Melted Save Files", "*.txt")]
        elif self.file_type.get() == "json":
            filetypes = [("JSON Files", "*.json")]
            
        files = filedialog.askopenfilenames(
            title="Select files",
            filetypes=filetypes
        )
        
        for file in files:
            if file not in self.selected_files:
                self.selected_files.append(file)
                self.file_listbox.insert(tk.END, os.path.basename(file))
                
    def remove_selected(self):
        selection = self.file_listbox.curselection()
        for index in reversed(selection):
            self.file_listbox.delete(index)
            self.selected_files.pop(index)
            
    def clear_files(self):
        self.file_listbox.delete(0, tk.END)
        self.selected_files.clear()
        
    def process_files(self):
        if not self.selected_files:
            messagebox.showwarning("No Files", "Please select files to process first.")
            return
            
        # Disable buttons during processing
        self.disable_buttons()
        
        # Start processing in a separate thread
        # thread = threading.Thread(target=self.process_files_thread)
        # thread.start()
        
    def process_files_thread(self):
        try:
            total = len(self.selected_files)
            all_data = []
            
            for i, file in enumerate(self.selected_files, 1):
                # Update progress
                progress = (i / total) * 100
                self.root.after(0, self.update_progress, f"Processing {os.path.basename(file)}", progress)
                
                # Process file based on type
                data = None
                if self.file_type.get() == "hoi4":
                    data = self.process_hoi4_save(file)
                elif self.file_type.get() == "txt":
                    data = self.process_melted_save(file)
                elif self.file_type.get() == "json":
                    data = self.process_json(file)
                
                if data:
                    all_data.append(data)
            
            if all_data:
                self.create_combined_excel(all_data)
                self.root.after(0, self.update_progress, f"Processed {len(all_data)} files successfully!", 100)
            else:
                self.root.after(0, self.update_progress, "No data was processed!", 100)
                
        finally:
            self.root.after(0, self.enable_buttons)
    
    def process_json(self, file):
        """Process JSON file and return analysis data."""
        import equipment
        import buildings
        import focus
        import intelligence
        from excel_generator import create_excel_report
        
        base_name = os.path.splitext(os.path.basename(file))[0]
        
        # Run all analyses
        equipment_data = equipment.analyze_save_file(file)
        buildings_data = buildings.analyze_save_file(file)
        focus_data = focus.analyze_save_file(file)
        intel_data = intelligence.analyze_save_file(file)
        
        # Generate Excel report if we have data
        if equipment_data and buildings_data:
            create_excel_report(equipment_data, buildings_data, base_name)
            # Combine all datasets into one
            return {
                'date': equipment_data['date'],
                'country_data': equipment_data['country_data'],
                'state_buildings': buildings_data['state_buildings'],
                'construction_queue': buildings_data['construction_queue'],
                'countries': {  # Add focus and intel data in correct structure
                    country_tag: {
                        'focus': focus_data.get('focus_data', {}).get(country_tag, {}),
                        'intelligence_agency': intel_data.get('agency_data', {}).get(country_tag, {})
                    }
                    for country_tag in equipment_data['country_data'].keys()
                }
            }
        return None

    def process_hoi4_save(self, file):
        """Process HOI4 save file and return analysis data."""
        from main import process_save_file
        base_name = os.path.splitext(os.path.basename(file))[0]
        json_path = os.path.join("melted_saves", f"{base_name}.json")
        if process_save_file(file):
            import equipment
            import buildings
            import focus
            import intelligence
            equipment_data = equipment.analyze_save_file(json_path)
            buildings_data = buildings.analyze_save_file(json_path)
            focus_data = focus.analyze_save_file(json_path)
            intel_data = intelligence.analyze_save_file(json_path)
            if equipment_data and buildings_data:
                return {
                    'date': equipment_data['date'],
                    'country_data': equipment_data['country_data'],
                    'state_buildings': buildings_data['state_buildings'],
                    'construction_queue': buildings_data['construction_queue'],
                    'countries': {  # Add focus and intel data in correct structure
                        country_tag: {
                            'focus': focus_data.get('focus_data', {}).get(country_tag, {}),
                            'intelligence_agency': intel_data.get('agency_data', {}).get(country_tag, {})
                        }
                        for country_tag in equipment_data['country_data'].keys()
                    }
                }
        return None

    def process_melted_save(self, file):
        """Process melted save file and return analysis data."""
        # Same update as process_hoi4_save
        from main import process_save_file
        base_name = os.path.splitext(os.path.basename(file))[0]
        json_path = os.path.join("melted_saves", f"{base_name}.json")
        if process_save_file(file):
            import equipment
            import buildings
            import focus
            import intelligence
            equipment_data = equipment.analyze_save_file(json_path)
            buildings_data = buildings.analyze_save_file(json_path)
            focus_data = focus.analyze_save_file(json_path)
            intel_data = intelligence.analyze_save_file(json_path)
            if equipment_data and buildings_data:
                return {
                    'date': equipment_data['date'],
                    'country_data': equipment_data['country_data'],
                    'state_buildings': buildings_data['state_buildings'],
                    'construction_queue': buildings_data['construction_queue'],
                    'countries': {  # Add focus and intel data in correct structure
                        country_tag: {
                            'focus': focus_data.get('focus_data', {}).get(country_tag, {}),
                            'intelligence_agency': intel_data.get('agency_data', {}).get(country_tag, {})
                        }
                        for country_tag in equipment_data['country_data'].keys()
                    }
                }
        return None

    def create_combined_excel(self, all_data):
        """Create combined Excel file from multiple analyses."""
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            major_countries = ['USA', 'ENG', 'GER', 'ITA', 'SOV', 'JAP']
            sheets_created = 0
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            section_font = Font(bold=True)

            # Custom function to convert any date format to sortable format
            def format_date_for_sort(date_str):
                if not date_str:
                    return ""
                    
                date_str = str(date_str).strip("' ")
                
                # Already in YYYY-MM-DD format
                if '-' in date_str and len(date_str) == 10:
                    return date_str
                    
                try:
                    # Try to parse various formats
                    if '/' in date_str:  # DD/MM/YYYY
                        day, month, year = map(int, date_str.split('/'))
                        return f"{year:04d}-{month:02d}-{day:02d}"
                    elif '-' in date_str:  # MMM-YY
                        month_str, year_str = date_str.split('-')
                        months = {'JAN':1, 'FEB':2, 'MAR':3, 'APR':4, 'MAY':5, 'JUN':6,
                                'JUL':7, 'AUG':8, 'SEP':9, 'OCT':10, 'NOV':11, 'DEC':12}
                        month = months[month_str.upper()]
                        year = 1900 + int(year_str) if int(year_str) > 50 else 2000 + int(year_str)
                        return f"{year:04d}-{month:02d}-01"
                except:
                    return date_str
                return date_str
            
            # Create sheets for each major country
            for country in major_countries:
                ws = wb.create_sheet(country)
                current_row = 1

                # 1. Production Section Header
                ws['A1'] = "Military Production"
                ws['A1'].font = section_font
                current_row += 2

                # Get equipment names specific to this country and map to types
                country_equipment = set()
                for data in all_data:
                    if country in data['country_data']:
                        # Map equipment names to their types using equipment_lines
                        country_data = data['country_data'][country]
                        for line in country_data.get('equipment_lines', []):
                            eq_name = line['equipment_name']
                            eq_type = EQ_TYPE.get(eq_name, eq_name)
                            country_equipment.add(eq_type)
                
                equipment_list = sorted(list(country_equipment))
                if not equipment_list:
                    continue  # Skip if country has no equipment
                
                # Setup headers using equipment types
                ws.cell(row=1, column=1, value='Date')
                cell = ws.cell(row=1, column=1)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
                
                for col, eq_type in enumerate(equipment_list, 2):
                    cell = ws.cell(row=1, column=col, value=eq_type)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.font = Font(color='FFFFFF', bold=True)
                
                # Add data rows mapping equipment names to types
                row = 2
                # Sort data by cleaned date string
                sorted_data = sorted(all_data, key=lambda x: format_date_for_sort(x.get('date', '')))
                for data in sorted_data:
                    if country in data['country_data']:
                        country_data = data['country_data'][country]
                        # Clean the date string before writing to cell
                        clean_date = data['date'].strip("' ") if data.get('date') else ''
                        ws.cell(row=row, column=1, value=clean_date)

                        # Create type-based factory counts from equipment lines
                        type_factories = defaultdict(int)
                        for line in country_data.get('equipment_lines', []):
                            eq_name = line['equipment_name']
                            eq_type = EQ_TYPE.get(eq_name, eq_name)
                            type_factories[eq_type] += line['active_factories']
                        
                        # Fill in factory counts by type
                        for col, eq_type in enumerate(equipment_list, 2):
                            factory_count = type_factories.get(eq_type, 0)
                            if factory_count > 0:
                                ws.cell(row=row, column=col, value=factory_count)
                        row += 1
                
                # After production section, add spacing
                row += 2

                # 2. Building Status Section
                ws.cell(row=row, column=1, value="Building Status")
                ws.cell(row=row, column=1).font = section_font
                row += 2

                # Get all dates and states for this country
                all_states = set()
                all_dates = []
                for data in all_data:
                    if 'state_buildings' in data:
                        all_dates.append(data['date'].strip("' "))
                        for state_id, buildings in data['state_buildings'].items():
                            if buildings.get('owner', '') == country:
                                all_states.add(state_id)

                # Sort states and dates
                sorted_states = sorted(all_states, key=lambda x: int(x) if x.isdigit() else float('inf'))
                all_dates = sorted(all_dates, key=lambda x: format_date_for_sort(x))

                # Add state headers
                for col, state_id in enumerate(sorted_states, 2):
                    cell = ws.cell(row=row, column=col, value=f"{state_id} - {STATES.get(state_id, 'Unknown')}")
                    cell.fill = header_fill
                    cell.font = header_font

                # Add date column header
                cell = ws.cell(row=row, column=1, value="Date")
                cell.fill = header_fill
                cell.font = header_font

                # Add building data rows
                row += 1
                for date in all_dates:
                    ws.cell(row=row, column=1, value=date)
                    data = next((d for d in all_data if d['date'].strip("' ") == date), None)
                    if data and 'state_buildings' in data:
                        for col, state_id in enumerate(sorted_states, 2):
                            if state_id in data['state_buildings']:
                                buildings = data['state_buildings'][state_id]
                                if buildings.get('owner', '') == country:
                                    # Format: infrastructure civs mils
                                    value = f"{buildings.get('infrastructure', 0)} {buildings.get('civs', 0)} {buildings.get('mils', 0)}"
                                    ws.cell(row=row, column=col, value=value)
                    row += 1

                row += 3

                # 3. Construction Queue Section
                ws.cell(row=row, column=1, value="Construction Queue")
                ws.cell(row=row, column=1).font = section_font
                row += 2

                # Add construction headers
                headers = ['Date', 'State', 'Building', 'Active Factories', 'Progress', 'Target', 'Started']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font

                # Add construction data
                row += 1
                for data in sorted_data:
                    if 'construction_queue' in data and country in data['construction_queue']:
                        for item in data['construction_queue'][country]:
                            ws.cell(row=row, column=1, value=data['date'].strip("' "))
                            ws.cell(row=row, column=2, value=f"{item['state']} - {item['state_name']}")
                            ws.cell(row=row, column=3, value=item['building'])
                            ws.cell(row=row, column=4, value=item['active_factories'])
                            ws.cell(row=row, column=5, value=item['produced'])
                            ws.cell(row=row, column=6, value=item['amount'])
                            ws.cell(row=row, column=7, value=item['created'])
                            row += 1

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width
                sheets_created += 1

            # If no sheets were created, add a default sheet
            if sheets_created == 0:
                ws = wb.create_sheet("Info")
                ws['A1'] = "No data found for major countries"
                ws['A1'].font = Font(bold=True)
                ws.column_dimensions['A'].width = 30

            # Save workbook
            excel_path = os.path.join('output', 'combined_analysis.xlsx')
            os.makedirs('output', exist_ok=True)
            wb.save(excel_path)
            print(f"Combined analysis saved to: {excel_path}")
            
        except Exception as e:
            print(f"Error creating combined Excel: {str(e)}")
            raise
    
    def update_progress(self, message, value):
        self.progress_var.set(message)
        self.progress_bar['value'] = value
        
    def disable_buttons(self):
        for widget in self.root.winfo_children():
            if isinstance(widget, ttk.Frame) or isinstance(widget, ttk.LabelFrame):
                for child in widget.winfo_children():
                    if isinstance(child, ttk.Button):
                        child['state'] = 'disabled'
                        
    def enable_buttons(self):
        for widget in self.root.winfo_children():
            if isinstance(widget, ttk.Frame) or isinstance(widget, ttk.LabelFrame):
                for child in widget.winfo_children():
                    if isinstance(child, ttk.Button):
                        child['state'] = 'normal'
    
    def generate_country_report(self):
        """Generate Excel report for a specific country."""
        tag = self.country_tag.get().strip().upper()
        if not tag:
            messagebox.showwarning("Invalid TAG", "Please enter a country TAG.")
            return
            
        if not self.selected_files:
            messagebox.showwarning("No Files", "Please select files to process first.")
            return
            
        # Disable buttons during processing
        self.disable_buttons()
        
        # Start processing in a separate thread
        thread = threading.Thread(target=lambda: self.process_country_report_thread(tag))
        thread.start()
        
    def process_country_report_thread(self, country_tag):
        """Process files and generate report for specific country in a separate thread."""
        try:
            total = len(self.selected_files)
            all_data = []
            
            for i, file in enumerate(self.selected_files, 1):
                # Update progress
                progress = (i / total) * 100
                self.root.after(0, self.update_progress, f"Processing {os.path.basename(file)}", progress)
                
                # Process file based on type
                data = None
                if self.file_type.get() == "hoi4":
                    data = self.process_hoi4_save(file)
                elif self.file_type.get() == "txt":
                    data = self.process_melted_save(file)
                elif self.file_type.get() == "json":
                    data = self.process_json(file)
                
                if data:
                    all_data.append(data)
            
            if all_data:
                # Create workbook
                wb = openpyxl.Workbook()
                wb.remove(wb.active)  # Remove default sheet
                
                # Create sheets
                production_ws = wb.create_sheet("Production")
                buildings_ws = wb.create_sheet("Buildings")
                construction_ws = wb.create_sheet("Construction")
                focus_ws = wb.create_sheet("Focus Completion")  # New sheet
                intel_ws = wb.create_sheet("Intelligence")  # New sheet
                
                # Use the same formatting
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                section_font = Font(bold=True)

                # Create all sheets
                self.create_production_sheet(production_ws, all_data, country_tag, header_fill, header_font)
                self.create_buildings_sheet(buildings_ws, all_data, country_tag, header_fill, header_font)
                self.create_construction_sheet(construction_ws, all_data, country_tag, header_fill, header_font)
                self.create_focus_sheet(focus_ws, all_data, country_tag, header_fill, header_font)
                self.create_intelligence_sheet(intel_ws, all_data, country_tag, header_fill, header_font)

                # Save workbook
                output_path = os.path.join('output', f'{country_tag}_analysis.xlsx')
                os.makedirs('output', exist_ok=True)
                wb.save(output_path)
                
                self.root.after(0, self.update_progress, 
                              f"Report generated for {country_tag}: {output_path}", 100)
                self.root.after(0, messagebox.showinfo, "Success", 
                              f"Report generated: {output_path}")
            else:
                self.root.after(0, self.update_progress, "No data was processed!", 100)
                
        except Exception as e:
            self.root.after(0, messagebox.showerror, "Error", 
                          f"Error generating report: {str(e)}")
        finally:
            self.root.after(0, self.enable_buttons)

    def create_production_sheet(self, ws, all_data, country_tag, header_fill, header_font):
        """Create the production sheet (existing functionality)"""
        current_row = 1
        ws['A1'] = "Military Production"
        ws['A1'].font = Font(bold=True)
        current_row += 2
        
        # Get equipment names for this country
        country_equipment = set()
        for data in all_data:
            if country_tag in data['country_data']:
                country_data = data['country_data'][country_tag]
                for line in country_data.get('equipment_lines', []):
                    eq_name = line['equipment_name']
                    eq_type = EQ_TYPE.get(eq_name, eq_name)
                    country_equipment.add(eq_type)
        
        if not country_equipment:
            self.root.after(0, messagebox.showinfo, "No Data", 
                          f"No production data found for country {country_tag}")
            return
        
        # Setup headers using equipment types
        equipment_list = sorted(list(country_equipment))
        headers = ['Date'] + equipment_list
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        current_row += 1
        
        # Add production data
        for data in sorted(all_data, key=lambda x: x.get('date', '')):
            if country_tag in data['country_data']:
                country_data = data['country_data'][country_tag]
                ws.cell(row=current_row, column=1, value=data['date'])
                
                # Create type-based factory counts
                type_factories = defaultdict(int)
                for line in country_data.get('equipment_lines', []):
                    eq_name = line['equipment_name']
                    eq_type = EQ_TYPE.get(eq_name, eq_name)
                    type_factories[eq_type] += line['active_factories']
                
                # Fill in factory counts by type
                for col, eq_type in enumerate(equipment_list, 2):
                    factory_count = type_factories.get(eq_type, 0)
                    if factory_count > 0:
                        ws.cell(row=current_row, column=col, value=factory_count)
                
                current_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_buildings_sheet(self, ws, all_data, country_tag, header_fill, header_font):
        """Create the buildings sheet with state columns and building data"""
        # Get all states owned by the country and all dates
        owned_states = {}  # Changed to dict to store state info
        dates = []
        
        for data in all_data:
            if 'date' in data:
                dates.append(data['date'].strip("' "))
            if 'state_buildings' in data:
                for state_id, buildings in data['state_buildings'].items():
                    if buildings.get('owner', '') == country_tag:
                        if state_id not in owned_states or owned_states[state_id] < buildings.get('manpower', 0):
                            owned_states[state_id] = buildings.get('manpower', 0)

        # Sort states by manpower (descending) and then by ID
        sorted_states = sorted(owned_states.keys(), 
                             key=lambda x: (-owned_states[x], int(x) if x.isdigit() else float('inf')))
        dates = sorted(dates)

        # Create headers
        current_col = 2  # Start from column B (A is for dates)
        ws.cell(row=1, column=1, value="Date")
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).font = header_font

        # Create merged state headers and sub-headers with alternating colors
        state_fill_1 = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        state_fill_2 = PatternFill(start_color='DEEAF6', end_color='DEEAF6', fill_type='solid')
        current_fill = state_fill_1

        for state_id in sorted_states:
            state_name = f"{state_id} - {STATES.get(state_id, 'Unknown')}"
            
            # Merge cells for state name
            ws.merge_cells(start_row=1, start_column=current_col, 
                          end_row=1, end_column=current_col + 2)
            cell = ws.cell(row=1, column=current_col, value=state_name)
            cell.fill = header_fill
            cell.font = header_font

            # Add sub-headers with state background color
            ws.cell(row=2, column=current_col, value="Infra").fill = current_fill
            ws.cell(row=2, column=current_col + 1, value="Civs").fill = current_fill
            ws.cell(row=2, column=current_col + 2, value="Mils").fill = current_fill
            
            # Add thin borders between states
            if current_col > 2:  # Not the first state
                for row in range(1, len(dates) + 3):  # +3 for headers
                    cell = ws.cell(row=row, column=current_col)
                    cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'))
            
            current_col += 3
            # Alternate background color for next state
            current_fill = state_fill_2 if current_fill == state_fill_1 else state_fill_1

        # Add data rows
        current_row = 3
        row_fill_1 = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        row_fill_2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        left_alignment = openpyxl.styles.Alignment(horizontal='left')

        for date_idx, date in enumerate(dates):
            # Set date cell alignment to left
            date_cell = ws.cell(row=current_row, column=1, value=date)
            date_cell.alignment = left_alignment
            
            current_col = 2
            data = next((d for d in all_data if d['date'].strip("' ") == date), None)
            row_fill = row_fill_1 if date_idx % 2 == 0 else row_fill_2
            
            if data and 'state_buildings' in data:
                for state_idx, state_id in enumerate(sorted_states):
                    state_fill = state_fill_1 if state_idx % 2 == 0 else state_fill_2
                    
                    if state_id in data['state_buildings']:
                        buildings = data['state_buildings'][state_id]
                        if buildings.get('owner', '') == country_tag:
                            for i in range(3):  # Three cells per state
                                cell = ws.cell(row=current_row, column=current_col + i)
                                cell.fill = state_fill
                                cell.alignment = left_alignment  # Set left alignment for each cell
                            
                            # Set values with left alignment
                            infra_cell = ws.cell(row=current_row, column=current_col, 
                                  value=buildings.get('infrastructure', 0))
                            civs_cell = ws.cell(row=current_row, column=current_col + 1, 
                                  value=buildings.get('civs', 0))
                            mils_cell = ws.cell(row=current_row, column=current_col + 2, 
                                  value=buildings.get('mils', 0))
                            
                            # Apply left alignment to all value cells
                            infra_cell.alignment = left_alignment
                            civs_cell.alignment = left_alignment
                            mils_cell.alignment = left_alignment
                            
                    current_col += 3
            
            current_row += 1

    def create_construction_sheet(self, ws, all_data, country_tag, header_fill, header_font):
        """Create the construction queue sheet organized by date"""
        # First collect all construction data by date
        construction_by_date = defaultdict(list)
        
        # Custom function to convert any date format to sortable format
        def format_date_for_sort(date_str):
            if not date_str:
                return ""
                
            date_str = str(date_str).strip("' ")
            
            # Already in YYYY-MM-DD format
            if '-' in date_str and len(date_str) == 10:
                return date_str
                
            try:
                # Try to parse various formats
                if '/' in date_str:  # DD/MM/YYYY
                    day, month, year = map(int, date_str.split('/'))
                    return f"{year:04d}-{month:02d}-{day:02d}"
                elif '.' in date_str:  # YYYY.MM.DD
                    year, month, day = map(int, date_str.split('.'))
                    return f"{year:04d}-{month:02d}-{day:02d}"
            except:
                return date_str
            return date_str
        
        # Get and sort all dates
        all_dates = set()
        for data in all_data:
            if 'construction_queue' in data and country_tag in data['construction_queue']:
                all_dates.add(data['date'].strip("' "))
        
        # Sort dates chronologically
        all_dates = sorted(all_dates, key=format_date_for_sort)
        
        # Process construction data
        for data in all_data:
            date = data['date'].strip("' ")
            if 'construction_queue' in data and country_tag in data['construction_queue']:
                for item in data['construction_queue'][country_tag]:
                    construction_by_date[date].append(item)

        # Set up the headers - each date gets 4 columns
        current_col = 1
        header_row = 1
        subheader_row = 2
        data_start_row = 3

        # Setup header structure
        headers = ['State', 'Building', 'Factories', 'Progress']
        cols_per_date = len(headers)  # Number of columns per date

        # Add date headers and subheaders
        for date in all_dates:
            if construction_by_date[date]:  # Only add date columns if there's data
                # Add vertical border between dates
                if current_col > 1:
                    for row in range(1, ws.max_row + 1):
                        cell = ws.cell(row=row, column=current_col)
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'))

                # Add merged date header
                ws.merge_cells(
                    start_row=header_row,
                    start_column=current_col,
                    end_row=header_row,
                    end_column=current_col + cols_per_date - 1
                )
                date_cell = ws.cell(row=header_row, column=current_col, value=date)
                date_cell.fill = header_fill
                date_cell.font = header_font
                date_cell.alignment = openpyxl.styles.Alignment(horizontal='center')

                # Add subheaders
                for i, header in enumerate(headers):
                    cell = ws.cell(row=subheader_row, column=current_col + i, value=header)
                    cell.fill = header_fill
                    cell.font = header_font

                # Add construction data
                data_row = data_start_row
                for item in construction_by_date[date]:
                    # State column
                    ws.cell(row=data_row, column=current_col,
                           value=f"{item['state']} - {item['state_name']}")
                    # Building column
                    ws.cell(row=data_row, column=current_col + 1,
                           value=item['building'])
                    # Factories column
                    ws.cell(row=data_row, column=current_col + 2,
                           value=item['active_factories'])
                    # Progress column
                    ws.cell(row=data_row, column=current_col + 3,
                           value=f"{item['produced']}/{item['amount']}")
                    data_row += 1

                current_col += cols_per_date

        # Auto-adjust column widths
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column = ws.column_dimensions[openpyxl.utils.get_column_letter(col)]
            
            for cell in ws[openpyxl.utils.get_column_letter(col)]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = max_length + 2
            column.width = adjusted_width

    def create_focus_sheet(self, ws, all_data, country_tag, header_fill, header_font):
        """Create sheet showing completed focuses"""
        # Setup header
        ws.cell(row=1, column=1, value="Completed Focuses")
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).font = header_font
        
        current_row = 2
        
        # Get country data from latest save
        latest_data = max(all_data, key=lambda x: x.get('date', ''))
        if 'countries' in latest_data and country_tag in latest_data['countries']:
            country_data = latest_data['countries'][country_tag]
            focus_data = country_data.get('focus', {})
            
            if isinstance(focus_data, dict):
                completed_focuses = focus_data.get('completed', [])
                # Add one row for each completed focus
                for focus_name in completed_focuses:
                    # Clean up focus name
                    clean_name = focus_name
                    if focus_name.startswith(country_tag + '_'):
                        clean_name = focus_name[len(country_tag)+1:]  # Remove TAG_
                    clean_name = clean_name.replace('_', ' ')  # Replace underscores with spaces
                    
                    ws.cell(row=current_row, column=1, value=clean_name)
                    current_row += 1
        
        # Auto-adjust column width
        ws.column_dimensions['A'].width = max(len(str(cell.value or "")) for cell in ws['A']) + 2

    def create_intelligence_sheet(self, ws, all_data, country_tag, header_fill, header_font):
        """Create sheet showing intelligence agency data by date"""
        
        # Define headers for the combined table
        headers = ["Date", "Total Slots", "Free Slots"] + [
            "Form Dept", "Crypto 1", "Crypto 2", "Suicide Pills", "Army Dept", 
            "Interrogation", "Defense", "Psycho War", "Air Dept", "Navy Dept",
            "Decrypt 1", "Training", "Decrypt 2", "Radios", "Explosives", "Diplo"
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Define upgrade mapping
        upgrade_mapping = {
            'upgrade_form_department': 4,
            'upgrade_crypto_strength': 5,
            'upgrade_crypto_strength_2': 6,
            'upgrade_suicide_pills': 7,
            'upgrade_army_department': 8,
            'upgrade_interrogation_techniques': 9,
            'upgrade_passive_defense': 10,
            'upgrade_psycho_warfare': 11,
            'upgrade_airforce_department': 12,
            'upgrade_naval_department': 13,
            'upgrade_decryption_boost': 14,
            'upgrade_training_centers': 15,
            'upgrade_decryption_boost_2': 16,
            'upgrade_portable_radios': 17,
            'upgrade_plastic_explosives': 18,
            'upgrade_diplo_training': 19
        }
        
        # Sort data by date
        sorted_data = sorted(all_data, key=lambda x: x.get('date', ''))
        current_row = 2
        
        # Process each save file
        for data in sorted_data:
            if 'countries' in data and country_tag in data['countries']:
                country_data = data['countries'][country_tag]
                if 'intelligence_agency' in country_data:
                    agency_data = country_data['intelligence_agency']
                    date = data.get('date', '')
                    
                    # Add all data in one row
                    ws.cell(row=current_row, column=1, value=date)
                    ws.cell(row=current_row, column=2, value=agency_data.get('max_operative_count', 0))
                    ws.cell(row=current_row, column=3, value=agency_data.get('usable_operative_slots', 0))
                    
                    # Add upgrade values
                    current_upgrades = agency_data.get('upgrades', {})
                    for upgrade_key, col in upgrade_mapping.items():
                        value = current_upgrades.get(upgrade_key, 0)
                        ws.cell(row=current_row, column=col, value=value)
                    
                    current_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
