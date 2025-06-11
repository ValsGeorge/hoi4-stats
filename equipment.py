import json
from read_with_pyradox import load_json_file
import os
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font

def get_equipment_details(data, equipment_id, equipment_type):
    """Find equipment details from the equipments section."""
    if 'equipments' in data:
        equipment_list = data['equipments']
        for eq_data in equipment_list:
            if isinstance(eq_data, dict):
                if eq_data.get('id') == equipment_id and eq_data.get('type') == equipment_type:
                    return eq_data
    return None

def get_equipment_name_from_registry(equipment_registry, eq_id, eq_type):
    """Find equipment name based on ID and type from the equipment registry."""
    # First try direct lookup from the equipment registry
    for eq_name, eq_data in equipment_registry.items():
        if isinstance(eq_data, list):  # Handle list of variants
            for variant in eq_data:
                if (variant.get('id') == eq_id and 
                    variant.get('type') == eq_type):
                    return eq_name
        elif isinstance(eq_data, dict):  # Handle single equipment entry
            if 'id' in eq_data:
                eq_info = eq_data['id']
                if (eq_info.get('id') == eq_id and 
                    eq_info.get('type') == eq_type):
                    return eq_name
            
    return f"Unknown Equipment (ID: {eq_id}, Type: {eq_type})"

def convert_time_to_string(time_obj):
    """Convert a pyradox Time object to string."""
    if hasattr(time_obj, 'to_python'):
        # Convert Time object to string representation
        date_parts = time_obj.to_python()
        if isinstance(date_parts, tuple):
            return f"{date_parts[0]}.{date_parts[1]}.{date_parts[2]}"
    elif hasattr(time_obj, '__str__'):
        return str(time_obj)
    return str(time_obj)

def analyze_save_file(json_path):
    """Analyze the entire save file for relevant data."""
    try:
        data = load_json_file(json_path)
        if not data:
            print("Failed to load JSON data")
            return

        # Extract and explicitly convert save date to string
        raw_date = data.get('date', 'Unknown Date')
        save_date = convert_time_to_string(raw_date)
        print(f"Debug: Converted date {raw_date} to {save_date}")
        
        # First analyze equipment to build the registry
        equipment_registry = analyze_equipment(data)
        # Then analyze organizations with the equipment registry
        analyze_organizations(data, equipment_registry)
        # Finally analyze production queue with date
        analyze_production_queue(data, equipment_registry, save_date)

    except Exception as e:
        print(f"Error analyzing save file: {str(e)}")
        raise

def analyze_equipment(data):
    """Analyze equipment data from the parsed save file."""
    try:
        print("\nAnalyzing equipment data...")
        
        # First, collect all equipment names and their details
        equipment = dict()
        if 'equipments' in data:
            for eq_name, eq_data in data['equipments'].items():
                equipment[eq_name] = []
                if isinstance(eq_data, list):
                    # Handle list of variants
                    for variant in eq_data:
                        if isinstance(variant, dict) and 'id' in variant:
                            equipment[eq_name].append({
                                'name': eq_name,
                                'id': variant['id'].get('id'),
                                'type': variant['id'].get('type'),
                                'creator': variant.get('creator', 'Unknown'),
                                'origin': variant.get('origin', '---'),
                                'ideas': variant.get('ideas', [])
                            })
                elif isinstance(eq_data, dict) and 'id' in eq_data:
                    # Handle single equipment entry
                    equipment[eq_name].append({
                        'name': eq_name,
                        'id': eq_data['id'].get('id'),
                        'type': eq_data['id'].get('type'),
                        'creator': eq_data.get('creator', 'Unknown'),
                        'origin': eq_data.get('origin', '---'),
                        'ideas': eq_data.get('ideas', [])
                    })
                            
            print(f"Found {len(equipment)} unique equipment types")

        # Write results to a text file
        output_path = os.path.join('output', 'equipment_analysis.txt')
        os.makedirs('output', exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("Equipment Analysis\n")
            f.write("=================\n\n")
            
            f.write("Equipment Types and Variants:\n")
            f.write("==========================\n\n")
            
            for eq_name, variants in sorted(equipment.items()):
                if variants:  # Only write if there are variants
                    f.write(f"{eq_name}\n")                
                    for variant in variants:
                        f.write(f"ID: {variant['id']}\n")
                        f.write(f"Type: {variant['type']}\n")
                        f.write(f"Creator: {variant['creator']}\n")
                        if variant['origin'] != '---':
                            f.write(f"Origin: {variant['origin']}\n")
                        if variant['ideas']:
                            f.write(f"Ideas: {variant['ideas']}\n")
                        f.write("\n")
        
        print(f"Equipment analysis has been written to: {output_path}")
        return equipment
            
    except Exception as e:
        print(f"Error analyzing equipment: {str(e)}")
        print("Data structure where error occurred:")
        print(json.dumps(data.get('equipments', {}), indent=2)[:500])
        raise

def analyze_organizations(data, equipment_registry):
    """Analyze organizations from the parsed save file."""
    try:
        print("\nAnalyzing production organizations...")
        org_registry = {}
        
        # Process country data
        if 'countries' in data:
            for country_tag, country_data in data['countries'].items():
                # Get production data for country
                if 'production' in country_data:
                    prod_data = country_data['production']
                    
                    # Look for industrial organizations
                    if 'industrial_organisations' in prod_data:
                        orgs_data = prod_data['industrial_organisations']
                        # print(f"\nFound {len(orgs_data)} organizations in {country_tag}")
                        
                        # Process each organization
                        for org_name, org_data in orgs_data.items():
                            if isinstance(org_data, dict):
                                org_details = {
                                    'name': org_name,
                                    'country': country_tag,
                                    'id': org_data.get('id', {}).get('id'),
                                    'type': org_data.get('id', {}).get('type'),
                                    'history': []
                                }
                                
                                # Get production history
                                if 'history' in org_data:
                                    for entry in org_data['history']:
                                        if isinstance(entry, dict):
                                            eq_data = entry.get('equipment', {})
                                            prod_data = entry.get('data', {})
                                            eq_id = eq_data.get('id')
                                            eq_type = eq_data.get('type')
                                            eq_name = get_equipment_name_from_registry(equipment_registry, eq_id, eq_type)
                                            org_details['history'].append({
                                                'equipment_name': eq_name,
                                                'equipment_id': eq_id,
                                                'equipment_type': eq_type,
                                                'units': prod_data.get('units', 0),
                                                'date': entry.get('date', 'Unknown')
                                            })
                                
                                org_registry[org_name] = org_details
                                # print(f"Added organization: {org_name}")

        # Write results to a text file in project directory
        output_path = os.path.join('output', 'organizations_analysis.txt')
        os.makedirs('output', exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("Industrial Organizations Analysis\n")
            f.write("==============================\n\n")
            
            # Group by country
            by_country = defaultdict(list)
            for org_name, details in org_registry.items():
                by_country[details['country']].append((org_name, details))
            
            # Write organized by country
            for country, orgs in sorted(by_country.items()):
                f.write(f"\nCountry: {country}\n")
                f.write("=" * 40 + "\n")
                
                for org_name, details in sorted(orgs):
                    f.write(f"\nOrganization: {org_name}\n")
                    f.write(f"ID: {details['id']}, Type: {details['type']}\n")
                    
                    if details['history']:
                        f.write("\nProduction History:\n")
                        for entry in details['history']:
                            f.write(f"  Equipment: {entry['equipment_name']} - [{entry['equipment_id']}/{entry['equipment_type']}]\n")
                            f.write(f"  Units: {entry['units']}\n")
                            f.write(f"  Date: {entry['date']}\n")
                            f.write("\n")
                    f.write("-" * 40 + "\n")

        print(f"Organizations analysis has been written to: {output_path}")
        return org_registry
            
    except Exception as e:
        print(f"Error analyzing organizations: {str(e)}")
        # Print the structure of a sample country for debugging
        if 'countries' in data and 'SOV' in data['countries']:
            print("\nExample SOV data structure:")
            print(json.dumps(data['countries']['SOV'].get('production', {}), indent=2)[:500])
        raise


def analyze_production_queue(data, equipment_registry, save_date):
    """Analyze the military production lines from the parsed save file."""
    try:
        print("\nAnalyzing production queue...")
        production_lines = []
        by_country = defaultdict(list)
        
        # Ensure save_date is a string
        save_date = convert_time_to_string(save_date)
        print(f"Debug: Using date string: {save_date}")
        
        if 'countries' in data:
            for country_tag, country_data in data['countries'].items():
                if 'production' in country_data:
                    prod_data = country_data['production']
                    if 'military_lines' in prod_data:
                        for line in prod_data['military_lines']:
                            if isinstance(line, dict):
                                eq_id = line.get('equipment_variant_index', {}).get('id')
                                eq_type = line.get('equipment_variant_index', {}).get('type')
                                eq_name = get_equipment_name_from_registry(equipment_registry, eq_id, eq_type)
                                
                                line_details = {
                                    'country': country_tag,
                                    'active_factories': line.get('active_factories', 0),
                                    'equipment_name': eq_name,
                                    'equipment_id': eq_id,
                                    'equipment_type': eq_type,
                                    'priority': line.get('priority', 0),
                                    'produced': line.get('produced', 0),
                                    'speed': line.get('speed', 0),
                                    'cost': line.get('cost', 0),
                                    'manufacturer_id': line.get('industrial_manufacturer', {}).get('id')
                                }
                                production_lines.append(line_details)
                                by_country[country_tag].append(line_details)  # Add to by_country here

        # Create Excel only if not being combined
        if not hasattr(analyze_production_queue, 'combining'):
            wb = openpyxl.Workbook()
        
            # Remove default sheet
            wb.remove(wb.active)
            
            # List of major countries to track
            major_countries = ['USA', 'ENG', 'GER', 'ITA', 'SOV', 'JAP']
            
            # Headers for each sheet
            headers = ['Date', 'Priority', 'Active Factories', 'Equipment', 'Equipment ID', 
                      'Equipment Type', 'Production Speed', 'Production Cost']
            
            # Create sheets for each major country
            for country in major_countries:
                if country in by_country:
                    ws = wb.create_sheet(country)
                    
                    # Add headers with formatting
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                        cell.font = Font(color='FFFFFF', bold=True)
                    
                    # Add data rows - save_date is now a string
                    row = 2
                    for line in sorted(by_country[country], key=lambda x: x['priority']):
                        ws.cell(row=row, column=1, value=save_date)  # Will work now as save_date is a string
                        ws.cell(row=row, column=2, value=line['priority'])
                        ws.cell(row=row, column=3, value=line['active_factories'])
                        ws.cell(row=row, column=4, value=line['equipment_name'])
                        ws.cell(row=row, column=5, value=line['equipment_id'])
                        ws.cell(row=row, column=6, value=line['equipment_type'])
                        ws.cell(row=row, column=7, value=round(line['speed'], 2))
                        ws.cell(row=row, column=8, value=round(line['cost'], 2))
                        row += 1
                    
                    # Auto-adjust column widths
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        ws.column_dimensions[column].width = adjusted_width

            # Save Excel file
            excel_path = os.path.join('output', 'production_analysis.xlsx')
            wb.save(excel_path)
            print(f"Production analysis has been written to: {excel_path}")

        # Return the data for potential combining
        return {
            'date': save_date,
            'by_country': by_country,
            'production_lines': production_lines
        }

    except Exception as e:
        print(f"Error analyzing production queue: {str(e)}")
        raise