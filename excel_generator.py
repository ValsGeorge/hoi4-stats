import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os
from collections import defaultdict
from state_defines import STATES
from eq_definitions import EQ_TYPE  # Add this import

def create_excel_report(equipment_data, buildings_data, base_name):
    """Generate Excel report with country-specific sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    major_countries = ['ENG', 'FRA', 'GER', 'USA', 'SOV', 'JAP', 'ITA']
    sheets_created = 0  # Track number of sheets created
    
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    section_font = Font(bold=True)
    
    for country in major_countries:
        if country not in equipment_data.get('country_data', {}) and country not in buildings_data.get('construction_queue', {}):
            continue
            
        ws = wb.create_sheet(country)
        sheets_created += 1  # Increment counter when sheet is created
        current_row = 1
        
        # Production Lines Section
        ws['A1'] = "Military Production"
        ws['A1'].font = section_font
        current_row += 2
        
        # Add production headers
        ws[f'A{current_row}'] = "Date"
        ws[f'A{current_row}'].fill = header_fill
        ws[f'A{current_row}'].font = header_font
        
        # Get equipment list for this country
        equipment_lines = []
        if country in equipment_data.get('country_data', {}):
            country_eq = equipment_data['country_data'][country]
            # Use equipment_lines instead of equipment_factories
            for line in country_eq.get('equipment_lines', []):
                if line['active_factories'] > 0:
                    equipment_lines.append({
                        'name': line['equipment_name'],
                        'factories': line['active_factories']
                    })
            
            # Add each production line as a separate row
            for line in equipment_lines:
                current_row += 1
                ws[f'A{current_row}'] = country_eq['date']
                ws[f'B{current_row}'] = line['name']
                ws[f'C{current_row}'] = line['factories']
        
        current_row += 3
        
        # Building Status Section
        ws[f'A{current_row}'] = "Building Status"
        ws[f'A{current_row}'].font = section_font
        current_row += 2
        
        # Add building headers
        headers = ['ID', 'State', 'Infrastructure', 'Civs', 'Mils', 'Steel Mills', 'Aluminium Mills', 'Synth Refineries', 'Fuel Silos', 'Airports']
        for col, header in enumerate(headers):
            col_letter = openpyxl.utils.get_column_letter(col + 1)
            ws[f'{col_letter}{current_row}'] = header
            ws[f'{col_letter}{current_row}'].fill = header_fill
            ws[f'{col_letter}{current_row}'].font = header_font
        
        # Add building data - filter by owner
        if 'state_buildings' in buildings_data:
            for state_id, buildings in sorted(buildings_data['state_buildings'].items()):
                # Only show states owned by this country that have factories
                if (buildings.get('owner', '') == country and 
                    (buildings.get('civs', 0) > 0 or buildings.get('mils', 0) > 0)):
                    current_row += 1
                    state_name = STATES.get(state_id, "Unknown")
                    ws[f'A{current_row}'] = f"{state_id}"
                    ws[f'B{current_row}'] = f"{state_name}"
                    ws[f'C{current_row}'] = buildings.get('infrastructure', 0)
                    ws[f'D{current_row}'] = buildings.get('civs', 0)
                    ws[f'E{current_row}'] = buildings.get('mils', 0)
                    ws[f'F{current_row}'] = buildings.get('steel_mills', 0)
                    ws[f'G{current_row}'] = buildings.get('aluminium_mills', 0)
                    ws[f'H{current_row}'] = buildings.get('synthetic_refinery', 0)
                    ws[f'I{current_row}'] = buildings.get('fuel_silo', 0)
                    ws[f'J{current_row}'] = buildings.get('air_base', 0)

        current_row += 3

        # Construction Queue Section
        ws[f'A{current_row}'] = "Construction Queue"
        ws[f'A{current_row}'].font = section_font
        current_row += 2
        
        # Add construction headers
        headers = ['State', 'Building', 'Active Factories', 'Progress/Target']
        for col, header in enumerate(headers):
            col_letter = openpyxl.utils.get_column_letter(col + 1)
            ws[f'{col_letter}{current_row}'] = header
            ws[f'{col_letter}{current_row}'].fill = header_fill
            ws[f'{col_letter}{current_row}'].font = header_font
        
        # Add construction data - only for owned states
        if 'construction_queue' in buildings_data and country in buildings_data['construction_queue']:
            for item in buildings_data['construction_queue'][country]:
                state_id = str(item['state'])
                # Check if state is owned by the country
                if state_id in buildings_data['state_buildings'] and buildings_data['state_buildings'][state_id].get('owner', '') == country:
                    current_row += 1
                    state_name = STATES.get(state_id, "Unknown")
                    ws[f'A{current_row}'] = f"{state_id} - {state_name}"
                    ws[f'B{current_row}'] = item['building']
                    ws[f'C{current_row}'] = item['active_factories']
                    ws[f'D{current_row}'] = f"{item['produced']}/{item['amount']}"
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # If no sheets were created, add a default sheet
    if sheets_created == 0:
        ws = wb.create_sheet("Info")
        ws['A1'] = "No data found for major countries"
        ws['A1'].font = Font(bold=True)
        ws.column_dimensions['A'].width = 30

    # Save workbook
    os.makedirs('output', exist_ok=True)
    output_path = os.path.join('output', f'{base_name}_analysis.xlsx')
    wb.save(output_path)
